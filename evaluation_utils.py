import numpy as np
import pandas as pd
from caserec.evaluation.item_recommendation import ItemRecommendationEvaluation
from openpyxl import load_workbook
from pygini import gini
from scipy.stats import entropy
from scipy.stats import ttest_rel
from scipy.stats import wilcoxon
from sklearn.preprocessing import MinMaxScaler


def evaluate(alg_name: str, prediction_file: str, train_file: str, test_file: str):
    """
    Evaluate the output file
    :param alg_name: name of the algorithm
    :param prediction_file: file of the predictions made by an algorithm
    :param train_file: train file path
    :param test_file: test file path
    :return: file on results folder with results
    """

    output_file = prediction_file.replace('outputs', 'results')
    accurate_metrics = ItemRecommendationEvaluation(sep=',').evaluate_with_files(prediction_file, test_file)
    if "ncf" not in prediction_file:
        diver_metrics = evaluate_diversity(prediction_file, train_file)
    else:
        diver_metrics = {}

    # print(diver_metrics)

    f = open(output_file, "w")
    f.write("--- " + alg_name + " --- \n")
    for key in accurate_metrics.keys():
        f.write(str(key) + " " + str(accurate_metrics[key]) + "\n")

    for key in diver_metrics.keys():
        f.write(str(key) + " " + str(diver_metrics[key]) + "\n")
    f.close()


def evaluate_diversity(prediction_file: str, train_file: str):
    """
    Evaluate the diversity of the recommenders considering the gini index, entropy and coverage of items
    :param prediction_file: output file path of the recommender
    :param train_file: train file path used by the recommender
    :return: the gini, entropy and coverage metrics for top 1, 3, 5 and 10 rankings
    """

    ats = [1, 3, 5, 10]
    train = pd.read_csv(train_file, header=None)
    out = pd.read_csv(prediction_file, header=None)
    out.columns = ['user_id', 'item_id', 'score']
    total_items = len(train[train.columns[1]].unique())
    div = {}

    # for all rankings, generate a dataset only with the ranking size considering the full output of the recommenders
    # get the probability of every item happening and calculate gini, entropy and coverage for each
    for at in ats:
        out_at = pd.DataFrame()
        out_u = out.set_index('user_id')
        for u in out_u.index.unique():
            out_at = pd.concat([out_at, out_u.loc[u][:at].reset_index()], ignore_index=True)

        out_at = out_at.set_index('item_id')
        out_at['count'] = out_at.index.value_counts()
        out_at['prob'] = out_at['count'] / len(out_at.index)
        out_at = out_at[~out_at.index.duplicated(keep='first')]
        l = out_at['prob'].to_list()
        probs = l + [0 for i in range(0, total_items - len(l))]

        g = gini(np.array(probs))
        en = entropy(probs, base=10)
        agg_div = len(l)
        cov = agg_div / total_items

        div['AGG_DIV@' + str(at)] = agg_div
        div['GINI@' + str(at)] = g
        div['ENTROPY@' + str(at)] = en
        div['COVERAGE@' + str(at)] = cov

    return div


def statistical_relevance(proposed: str, baseline: str, dataset: str, metrics: list, method='both', save=False):
    """
    Function that calculates the statistical relevance with p_value
    :param proposed: proposed method final name. e.g.: path[policy=last_items=01_reorder=10]
    :param baseline: baseline name e.g.: bprmf, knn, mostpop, wikidata_page_rank8020
    :param dataset: path to the folds of the chosen dataset. e.g. "./datasets/ml-latest-small/folds/"
    :param metrics: list with the methods do analyse from the list
                    ["PREC, RECALL", "MAP", "NDCG", "GINI", "ENTROPY", "AGG_DIV", "COVERAGE"]
    :param method: method to test the statistical relevance, either 'ttest', 'wilcoxon' or 'both' that is the default value
    :param save: flag to save file in dataset directory
    :return: the statistical relevance of the proposed with the baseline for the metrics chosen for @1, @3, @5 and @10
    """
    div_metrics = ["GINI", "ENTROPY", "COVERAGE", "AGG_DIV"]
    results = pd.DataFrame(columns=['METRIC',
                                    'PROPOSED NAME', 'PROPOSED MEAN',
                                    'BASELINE NAME', 'BASELINE MEAN',
                                    'WILCOXON', 'TTEST'])

    ats = [1, 3, 5, 10]
    base_results = {}
    for m in metrics:
        for at in ats:
            base_results[m + "@" + str(at)] = []

    prop_results = {}
    for m in metrics:
        for at in ats:
            prop_results[m + "@" + str(at)] = []

    for i in range(0, 10):
        base_path = dataset + str(i) + "/results/" + baseline + ".csv"
        prop_path = dataset + str(i) + "/results/" + baseline + "_lodreorder_" + proposed + ".csv"

        try:
            base_df = file_to_df(base_path, baseline)
            prop_df = file_to_df(prop_path, proposed)
        except FileNotFoundError:
            continue

        for m in metrics:
            for at in ats:
                base_value = float(base_df[(base_df['metric'] == m) & (base_df['@'] == at)]['value'])
                base_results[m + "@" + str(at)].append(base_value)

                prop_value = float(prop_df[(prop_df['metric'] == m) & (prop_df['@'] == at)]['value'])
                prop_results[m + "@" + str(at)].append(prop_value)

    for m in metrics:
        for at in ats:

            if ("reorder=10" in proposed) and (m in div_metrics) and (at == 10):
                continue

            key = m + "@" + str(at)
            print("---" + key + "---")

            base_list = base_results[key]
            base_mean = sum(base_list) / len(base_list)
            print("Results of the baseline algorithm: " + str(base_list) +
                  " mean: " + str(base_mean) + " -> " + baseline)

            prop_list = prop_results[key]
            prop_mean = sum(prop_list) / len(prop_list)
            print("Results of the proposed algorithm: " + str(prop_list) +
                  " mean: " + str(prop_mean) + " -> " + proposed)

            wp = None
            tp = None
            if method == 'ttest':
                tt, tp = ttest_rel(base_list, prop_list)
                print("p-value with t-test: " + str(tp))
            elif method == 'wilcoxon':
                wt, wp = wilcoxon(base_list, prop_list)
                print("p-value with wilcoxon: " + str(wp))
            else:
                wt, wp = wilcoxon(base_list, prop_list)
                print("p-value with wilcoxon: " + str(wp))
                tt, tp = ttest_rel(base_list, prop_list)
                print("p-value with t-test: " + str(tp))

            results = results.append({'METRIC': key,
                                      'PROPOSED NAME': proposed, 'PROPOSED MEAN': prop_mean,
                                      'BASELINE NAME': base_mean, 'BASELINE MEAN': baseline,
                                      'WILCOXON': wp, 'TTEST': tp}, ignore_index=True)

    if save:
        p = dataset[:-6] + "results_" + proposed + ".xlsx"
        try:
            book = load_workbook(p)
            writer = pd.ExcelWriter(p, mode='r+')
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        except FileNotFoundError:
            writer = pd.ExcelWriter(p, mode='w+')
        results.to_excel(writer, sheet_name=baseline, index=False)
        writer.save()
        writer.close()
        print("--- FILE SAVE AT " + p + "---")


def file_to_df(file_path: str, algorithm: str):
    """
    Transform a result file into
    :param file_path: path to the result file of the algorithm
    :param algorithm: name of the algorithm
    :return: pandas df with the results
    """
    df = pd.DataFrame()
    f = open(file_path, 'r')
    for l in f.readlines()[1:]:
        l = l.split("\n")[0]
        metric_value = l.split('@')
        if metric_value[0] != metric_value[-1]:
            metric = metric_value[0]
            at_value = metric_value[-1].split(" ")
            at = int(at_value[0])
            value = at_value[-1]
            df = df.append({'alg': algorithm, '@': at, 'metric': metric, 'value': value}, ignore_index=True)
    f.close()
    return df


def evaluate_explanations(file_name: str, m_items: list, m_props: list, total_items: dict, total_props: dict,
                          train_set: pd.DataFrame, prop_set: pd.DataFrame, all_lir: list, all_sep: list, all_etd: list):
    """
    Diversity of explanation metrics. We will use 10 metrics: mean and std item and prop user diversity
    total item and prop aggregate diversity and entropy and gini
    :param file_name: the name of the results file
    :param m_items: distribution of different historic items shown to user in explanations
    :param m_props:distribution of different KG props shown to user in explanations
    :param total_items: dict of items, quantity of times used in explanations
    :param total_props: dict of prop, quantity of times used in explanations
    :param train_set: pandas dataframe of the train set
    :param prop_set: pandas dataframe of the properties extracted from the KG
    :param all_etd: list of all users etd metric
    :param all_sep: list of all users sep metric
    :param all_lir: list of all users lir metric
    :return: file saved with metrics
    """
    t_items = len(train_set[train_set.columns[0]].unique())
    t_props = len(prop_set[prop_set.columns[-1]].unique())
    items_distr = list(total_items.values())
    props_distr = list(total_props.values())
    item_probs = items_distr + [0 for i in range(0, t_items - len(items_distr))]
    props_probs = props_distr + [0 for i in range(0, t_props - len(props_distr))]

    mean_useritem_aggr = "Mean user item aggregate diversity: " + str(np.array(m_items).mean())
    std_useritem_aggr = "Std user item aggregate diversity: " + str(np.array(m_items).std())
    mean_userprop_aggr = "Mean user property aggregate diversity: " + str(np.array(m_props).mean())
    std_userprop_aggr = "Std user property aggregate diversity: " + str(np.array(m_props).std())
    total_items_str = "Total items aggregate diversity: " + str(len(total_items))
    total_props_str = "Total property aggregate diversity: " + str(len(total_props))
    ientropy = "Items entropy: " + str(entropy(item_probs))
    pentropy = "Props entropy: " + str(entropy(props_probs))
    igini = "Items Gini index: " + str(gini(np.array(item_probs, dtype=np.float64)))
    pgini = "Props Gini index: " + str(gini(np.array(props_probs, dtype=np.float64)))
    mean_lir = "LIR metric: " + str(np.array(all_lir).mean())
    mean_etd = "ETD metric: " + str(np.array(all_etd).mean())
    mean_sep = "SEP metric: " + str(np.array(all_sep).mean())
    std_lir = "std LIR metric: " + str(np.array(all_lir).std())
    std_etd = "std ETD metric: " + str(np.array(all_etd).std())
    std_sep = "std SEP metric: " + str(np.array(all_sep).std())

    f = open(file_name, mode="w", encoding='utf-8')
    f.write(file_name + "\n")
    f.write(mean_useritem_aggr + "\n")
    f.write(std_useritem_aggr + "\n")
    f.write(mean_userprop_aggr + "\n")
    f.write(std_userprop_aggr + "\n")
    f.write(total_items_str + "\n")
    f.write(total_props_str + "\n")
    f.write(ientropy + "\n")
    f.write(pentropy + "\n")
    f.write(igini + "\n")
    f.write(pgini + "\n")
    f.write(mean_lir + "\n")
    f.write(mean_etd + "\n")
    f.write(mean_sep + "\n")
    f.write(std_lir + "\n")
    f.write(std_etd + "\n")
    f.write(std_sep + "\n")
    f.close()

    print("\n" + file_name)
    print(mean_useritem_aggr)
    print(std_useritem_aggr)
    print(mean_userprop_aggr)
    print(std_userprop_aggr)
    print(total_items_str)
    print(total_props_str)
    print(ientropy)
    print(pentropy)
    print(igini)
    print(pgini)
    print(mean_lir)
    print(mean_etd)
    print(mean_sep)
    print(std_lir)
    print(std_etd)
    print(std_sep)


def lir_metric(beta: float, user: int, items: list, train_set: pd.DataFrame):
    """
    Linking Interaction Recency (LIR): metric proposed in https://dl.acm.org/doi/abs/10.1145/3477495.3532041
    :param beta: parameter for the exponential decay
    :param user: user id
    :param items: list  of listof items used for each recommendation explanation path
    :param train_set: training set of user item interactions
    :return: the LIR metric for the user, the lir for every recommendation is the mean of the lir of every recommendation
        and the lir for every recommendation is the mean of the lir for every item in the explanation path
    """
    # get user data
    if train_set.columns[0] == "artist_id":
        train_set = pd.read_csv("./datasets/hetrec2011-lastfm-2k/user_taggedartists-timestamps.dat", sep="\t")
        train_set = train_set[["userID", "artistID", "timestamp"]].set_index("userID")
    interacted = train_set.loc[user]
    interacted = interacted.reset_index()
    last_col = interacted.columns[-1]
    interacted = interacted.sort_values(last_col, ascending=True)
    interacted["lir"] = -1

    # for every item calculate the exponential decay
    min = interacted[last_col].min()
    last_value = min
    last_lir = min
    for i, row in interacted.iterrows():
        # if it is min, then lir is the value
        if row[last_col] == min:
            interacted.at[i, "lir"] = min
        # else if the count is the same repeat the sep, otherwise, calculate new sep
        else:
            if row[last_col] == last_value:
                interacted.at[i, "lir"] = last_lir
            else:
                lir = (1 - beta) * last_lir + beta * row[last_col]
                interacted.at[i, "lir"] = int(lir)
                last_value = row[last_col]
                last_lir = lir

    scaler = MinMaxScaler()
    interacted['normalized'] = scaler.fit_transform(
        np.asarray(interacted[interacted.columns[-1]]).astype(np.float64).reshape(-1, 1)).reshape(-1)

    # initialize mean variables
    total_sum = 0
    total_n = 0
    for pro_list in items:
        items_sum = 0
        items_n = 0
        for item in pro_list:
            try:
                value = int(interacted[interacted[interacted.columns[1]] == item]['normalized'])
                items_sum = items_sum + value
                items_n = items_n + 1
            except TypeError:
                pass

        try:
            total_sum = total_sum + (items_sum / items_n)
            total_n = total_n + 1
        except ZeroDivisionError:
            total_n = total_n + 1

    return total_sum / total_n


def sep_metric(beta: float, props: list, prop_set: pd.DataFrame, memo_sep: dict):
    """
    Shared Entity Popularity (SEP) metric proposed in https://dl.acm.org/doi/abs/10.1145/3477495.3532041
    :param beta: parameter for the exponential decay
    :param props: list of list of properties used for each recommendation explanation path
    :param prop_set: property set extrated from Wikidata
    :param memo_sep: memoization for sep values across users
    :return: the sep metric for the user, the sep for every recommendation is the mean of the sep of every recommendation
        and the sep for every recommendation is the mean of the sep for every item in the explanation path
    """

    # user variables for the mean sep of each explanation and scaler
    total_sum = 0
    total_n = 0
    scaler = MinMaxScaler()
    # for every list of properties in the user list of explanations
    for expl_props in props:
        # explanation variables for the mean sep of each explanation
        items_sum = 0
        items_n = 0
        # for every property list of each explanation
        for p in expl_props:
            # obtain the most popular link to of the property e.g. link actor from property Brad Pitt
            links = list(set(prop_set[prop_set["obj"] == p]['prop'].values))
            l_memo = list(set(memo_sep.keys()).intersection(set(links)))
            if len(l_memo) > 0:
                memo_df = memo_sep[l_memo[0]]
                p_sep_value = memo_df.loc[p][-1]
            else:
                link_df = prop_set[prop_set["prop"].isin(links)]
                # generate dataset with property as index and count as column
                count_link = pd.DataFrame(link_df.groupby("obj").count())
                count_link = count_link.sort_values(by=count_link.columns[0], ascending=True)
                count_link = pd.DataFrame(count_link[count_link.columns[0]])
                # initialize sep column with value -1
                count_link["sep"] = -1

                # obtain min value so we do not need to calculate every time
                # and initialize the last value and last sep as min according to the base case
                min = count_link[count_link.columns[0]].min()
                last_value = min
                last_sep = min
                for i, row in count_link.iterrows():
                    # if it is min, then lir is the value
                    if row[0] == min:
                        count_link.at[i, "sep"] = min
                    # else if the count is the same repeat the sep, otherwise, calculate new sep
                    else:
                        if row[0] == last_value:
                            count_link.at[i, "sep"] = last_sep
                        else:
                            sep = (1 - beta) * last_sep + beta * row[0]
                            count_link.at[i, "sep"] = sep
                            last_value = row[0]
                            last_sep = sep

                # generate normalized sep column
                try:
                    count_link['normalized'] = scaler.fit_transform(
                        np.asarray(count_link[count_link.columns[-1]]).astype(np.float64).reshape(-1, 1)).reshape(-1)
                except ValueError:
                    continue
                p_sep_value = count_link.loc[p][-1]
                for l in links:
                    memo_sep[l] = count_link

            # obtain sep value for the property and calculate mean
            items_sum = items_sum + p_sep_value
            items_n = items_n + 1

        # calculate total mean
        try:
            total_sum = total_sum + (items_sum / items_n)
            total_n = total_n + 1
        except ZeroDivisionError:
            total_n = total_n + 1

    return total_sum / total_n


def etd_metric(explanation_types: list, k: int, total_types: int):
    """
    Metric proposed by Ballocu 2022
    :param explanation_types: list of explanation types used in the explanations
    :param k: number of recommendations
    :param total_types: total number of explanation types in the dataset
    :return: the division between the explanation types in the explanations and the minimum between the k and total_types
    """
    return len(set(explanation_types)) / (min(k, total_types))


def explanation_file_to_df(file_path: str, algorithm: str):
    """
    Function that generates a dataframe with the offline explanations metrics for the algorithm
    "algorithm" on file "file path"
    :param file_path: path of the file
    :param algorithm: name of the algorithm
    :return: dataframe with offline explanation metrics of the algorithm retrieved from file
    """
    df = pd.DataFrame()
    f = open(file_path, 'r')
    for l in f.readlines()[1:]:
        l = l.split("\n")[0]
        metric_value = l.split(':')
        if metric_value[0] != metric_value[-1]:
            metric = metric_value[0]
            at_value = metric_value[1].split(" ")
            value = at_value[-1]
            df = df.append({'alg': algorithm, 'metric': metric, 'value': value}, ignore_index=True)
            # print(str(metric) + " " + str(value))
    f.close()
    return df


def statistical_relevance_explanations(rec_alg: str, dataset: str, reordered: int, n_explain: int):
    """
    Compute statistical relevance test for explanations evaluation
    :param rec_alg: recommendation algorithm used
    :param dataset: dataset (ml or last-fm)
    :param reordered: 1 if the recommendations were reordred by proposed reordering system 0 if not
    :param n_explain: number of explanations explained in the file to check for statistical significance
    :return: excel file with wilcoxon and ttest statistical relevance tests for all metrics
    """
    explod_results = []
    explodv2_results = []
    pem_results = []

    path = "./datasets/"
    if dataset == "ml":
        path = path + "ml-latest-small"
    else:
        path = path + "hetrec2011-lastfm-2k"
    path_base = path + "/folds/"

    for i in range(0, 10):
        path = path_base + str(i) + "/results/explanations/" + "reordered_recs=" + str(reordered)

        n_explain_s = "_"
        if n_explain != 5:
            n_explain_s = n_explain_s + "n_explain=" + str(n_explain) + "_"

        path_explod = path + "_expl_alg=explod" + n_explain_s + str(rec_alg) + ".csv"
        path_explod_v2 = path + "_expl_alg=explod_v2" + n_explain_s + str(rec_alg) + ".csv"
        path_pem = path + "_expl_alg=pem" + n_explain_s + str(rec_alg) + ".csv"

        explod_results.append(explanation_file_to_df(path_explod, "explod").set_index("metric"))
        explodv2_results.append(explanation_file_to_df(path_explod_v2, "pem").set_index("metric"))
        pem_results.append(explanation_file_to_df(path_pem, "explod_v2").set_index("metric"))

    metrics = list(explod_results[0].index.unique())
    results_df = pd.DataFrame(columns=["METRIC",
                                       "VERSION1", "VERSION1_MEAN",
                                       "VERSION2", "VERSION2_MEAN",
                                       "VERSION3", "VERSION3_MEAN",
                                       "WILCOXON12", "WILCOXON13", "WILCOXON23"])
    for m in metrics:
        m_explod = []
        m_explodv2 = []
        m_pem = []
        for i in range(0, 10):
            m_explod.append(float(explod_results[i].loc[m].value))
            m_explodv2.append(float(explodv2_results[i].loc[m].value))
            m_pem.append(float(pem_results[i].loc[m].value))

        explod_mean = np.array(m_explod).mean()
        print(m + " baseline results: ", m_explod)

        explodv2_mean = np.array(m_explodv2).mean()
        print(m + " proposed results: ", m_explodv2)

        pem_mean = np.array(m_pem).mean()
        print(m + " proposed results: ", m_pem)

        try:
            wt_12, wp_12 = wilcoxon(m_explod, m_explodv2)
            wt_13, wp_13 = wilcoxon(m_explod, m_pem)
            wt_23, wp_23 = wilcoxon(m_explodv2, m_pem)
        except ValueError:
            wt_12, wp_12 = 0, 0
            wt_13, wp_13 = 0, 0
            wt_23, wp_23 = 0, 0

        results_df = results_df.append({"METRIC": m,
                                        "VERSION1": "explod",
                                        "VERSION1_MEAN": explod_mean,
                                        "VERSION2": "explod_v2",
                                        "VERSION2_MEAN": explodv2_mean,
                                        "VERSION3": "pem",
                                        "VERSION3_MEAN": pem_mean,
                                        "WILCOXON12": wp_12, "WILCOXON13": wp_13, "WILCOXON23": wp_23
                                        }, ignore_index=True)

    p = path_base[:-6] + "explanation_results_reordered=" + str(reordered) + "_"
    if n_explain != 5:
        p = p + "n_explain=" + str(n_explain)
    p = p + ".xlsx"

    try:
        book = load_workbook(p)
        writer = pd.ExcelWriter(p, mode='r+')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    except FileNotFoundError:
        writer = pd.ExcelWriter(p, mode='w+')
    results_df.to_excel(writer, sheet_name=rec_alg, index=False)
    writer.save()
    writer.close()
    print("--- FILE SAVE AT " + p + "---")


def maut(dataset: str, folder: int, expl_algs: str, rec_alg: str, metrics: str, weights: str, n_explain: int):
    metrics_dict = {
        "SEP": "SEP metric",
        "ETD": "ETD metric",
        "LIR": "LIR metric",
        "TID": "Total items aggregate diversity",
        "TPD": "Total property aggregate diversity",
        "MID": "Mean user item aggregate diversity",
        "MPD": "Mean user property aggregate diversity",
        "IEnt": "Items entropy",
        "PEnt": "Props entropy",
        "IGini": "Items Gini index",
        "PGini": "Props Gini index"
    }
    expl_algs_l = expl_algs.split(" ")

    metrics_l = metrics.split(" ")
    if len(metrics_l) == 1:
        metrics_l = list(metrics_dict.keys())

    weights_l = weights.split(" ")
    if len(weights_l) == 1:
        weights_l = [1 / len(metrics_l) for _ in metrics_l]

    weights_dict = {}
    for i in range(0, len(metrics_l)):
        metric = metrics_l[i]
        weight = weights_l[i]
        weights_dict[metrics_dict[metric]] = weight

    path = "./datasets/"
    if dataset == "ml":
        path = path + "ml-latest-small"
    else:
        path = path + "hetrec2011-lastfm-2k"
    path_base = path + "/folds/"
    path = path_base + str(folder) + "/results/explanations/" + "reordered_recs=0"
    if n_explain != 5:
        path_n_expain = "n_explain=" + str(n_explain)
    df_metrics = []
    for i in range(0, len(expl_algs_l)):
        alg = expl_algs_l[i]
        if alg.startswith("webmedia"):
            path = path.replace("/explanations/", "/explanations/webmedia/")
        elif alg.startswith("llm"):
            path = path.replace("/explanations/", "/explanations/llm/")

        if n_explain != 5:
            alg_path = path + "_expl_alg=" + alg + "_" + path_n_expain + "_" + str(rec_alg) + ".csv"
        else:
            alg_path = path + "_expl_alg=" + alg + "_" + str(rec_alg) + ".csv"
        df_metrics.append(explanation_file_to_df(alg_path, alg))

        if alg.startswith("llm") or alg.startswith("webmedia"):
            path = path_base + str(folder) + "/results/explanations/" + "reordered_recs=0"

    scaler = MinMaxScaler()
    utility_matrix = pd.concat(df_metrics).pivot(index='alg', columns='metric', values='value').astype(np.float64)

    if "IGini" in metrics_l:
        utility_matrix[metrics_dict['IGini']] = 1 - utility_matrix[metrics_dict['IGini']]
    if "PGini" in metrics_l:
        utility_matrix[metrics_dict['PGini']] = 1 - utility_matrix[metrics_dict['PGini']]

    used_metrics = [metrics_dict[m] for m in metrics_l]

    utility_matrix_print = utility_matrix[used_metrics + ["std " + x for x in used_metrics]].copy()
    for m in used_metrics:
        truncated_values = utility_matrix_print[m].map(lambda x: "{:.4f}".format(x))
        std_values = utility_matrix_print['std ' + m].map(lambda x: "{:.2f}".format(x))
        utility_matrix_print[m] = truncated_values + "#" + std_values

    utility_matrix_print = utility_matrix_print[used_metrics]
    print(utility_matrix_print.T)
    print()

    utility_matrix = utility_matrix[used_metrics]
    uf_utility_matrix = pd.DataFrame(scaler.fit_transform(utility_matrix), columns=utility_matrix.columns).set_index(
        utility_matrix.index)

    for m in weights_dict.keys():
        uf_utility_matrix[m] = uf_utility_matrix[m] * float(weights_dict[m])

    rank = uf_utility_matrix.sum(axis=1).sort_values(ascending=False)
    for ind in rank.index:
        print(str(ind) + ": " + str(rank[ind]))
